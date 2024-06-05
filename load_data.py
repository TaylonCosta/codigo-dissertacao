import yaml
import pandas
import argparse
from pulp import *
from openpyxl import load_workbook


class Load_data:
    def ler_cenario(self, nome_arquivo):
        "Abre o arquivo YAML com parâmetros do cenário"
        with open(nome_arquivo, 'r') as arquivo:
            cenario = yaml.safe_load(arquivo)
        return cenario


    def load(self):
        parser = argparse.ArgumentParser(description='Otimizador Plano Semanal')
        parser.add_argument('-c', '--cenario', default='cenarios/ws1.yaml', type=str, help='Caminho para o arquivo do cenário a ser experimentado')
        parser.add_argument('-s', '--solver', default='GUROBI', type=str, help='Nome do otimizador a ser usado')
        parser.add_argument('-o', '--pasta-saida', default='experimentos', type=str, help='Pasta onde serão salvos os arquivos de resultados')
        parser.add_argument('--relax-and-fix', action='store_true', help='Habilita a heurística Relax And Fix das variáveis do mineroduto')
        parser.add_argument('--opt-partes', action='store_true', help='Habilita a heurística de otimização por partes')
        parser.add_argument('--ppo', action='store_true', help='Resolve o mdoelo pelo ppo')

        args = parser.parse_args()

        print(f'[OK]\nLendo arquivo {args.cenario}...   ', end='')
        # Abre o arquivo YAML com dados do cenário (parâmetros do problema)
        cenario = self.ler_cenario(args.cenario)

        # -----------------------------------------------------------------------------
        # Solver
        print(f'[OK]\nInstanciando solver {args.solver}...   ', end='')
        solver = getSolver(args.solver, msg=True)
        # solver.options.remove(("SolutionLimit", value))
        #------------------------------------------------------------------------------

        print(f'[OK]\nAbrindo planilha {cenario["geral"]["planilha"]}...   ', end='')
        # Abre a planilha e força o cálculo das fórmulas
        wb = load_workbook(cenario['geral']['planilha'], data_only=True)
        wb.calculation.calcMode = "auto"

        # -----------------------------------------------------------------------------
        # Índices usados pelas variáveis e parâmetros

        print(f'[OK]\nCriando índices...   ', end='')
        dias =      [f'd{dia+1:02d}' for dia in range(7)]    # d01, d02, ...   , d14
        horas =     [f'h{hora+1:02d}' for hora in range(24)]  # h01, h02, ...   , h6
        horas_D14 = [f'{dia}_{hora}' for dia in dias for hora in horas] # d01_h01, d01_h02, ...   , d14_h6
        horas_Dm3 = [f'dm{dia+1:02d}_h{hora+1:02d}' for dia in range(-4,-1) for hora in range(len(horas))] # dm-3_h01, dm-3_h02, ...   , dm-1_h6
        horas_Dm3_D14 = horas_Dm3 + horas_D14
        produtos_conc = cenario['concentrador']['produtos_conc']
        produtos_usina = cenario['usina']['produtos_usina']

        de_para_produtos_mina_conc = {'PRDT1': {'PRDT_C1': 1, 'PRDT_C2': 1, 'PRDT_C3': 1},
                                      'PRDT2': {'PRDT_C1': 0, 'PRDT_C2': 0, 'PRDT_C3': 1}}
        de_para_produtos_conc_usina = {'PRDT_C1': {'PRDT_U1': 0, 'PRDT_U2':0, 'PRDT_U3':0, 'PRDT_U4':1},
                                    'PRDT_C2': {'PRDT_U1': 0, 'PRDT_U2':1, 'PRDT_U3':1, 'PRDT_U4':0},
                                    'PRDT_C3': {'PRDT_U1': 1, 'PRDT_U2':0, 'PRDT_U3':0, 'PRDT_U4':0}}

        # Obs.: índices dos navios são definidos ao ler os dados da aba NAVIOS

        BIG_M = 10e6 # Big M

        opcoes_restricoes = ['fixado_pelo_usuario', 'fixado_pelo_modelo', 'livre']

        print(f'[OK]\nObtendo parâmetros do cenário...   ', end='')

        janela_planejamento = cenario['geral']['janela_planejamento']

        #verificar se não vai mesmo britagem

        taxa_alimentacao_britagem = cenario['mina']['taxa_alimentacao_britagem']
        disponibilidade_britagem = cenario['mina']['disponibilidade_britagem']
        utilizacao_britagem = cenario['mina']['utilizacao_britagem']

        taxa_producao_britagem = {dias[d]: taxa_alimentacao_britagem[d] *
                                        disponibilidade_britagem[d]/100 *
                                        utilizacao_britagem[d]/100
                                for d in range(len(dias))}
        print('--------------------------')
        print(taxa_producao_britagem)
        # Mina
        campanha_c3 = cenario['mina']['campanha']
        produtos_mina = set()
        for campanha in campanha_c3:
            produtos_mina.add(campanha[0])
        produtos_mina = list(produtos_mina)

        produtos_britagem = {produto: {hora: 0 for hora in horas_D14} for produto in produtos_mina}
        campanha_atual = 0
        for hora in horas_D14:
            if campanha_atual < len(campanha_c3) and hora == campanha_c3[campanha_atual][1]:
                produto_atual = campanha_c3[campanha_atual][0]
                campanha_atual += 1
            produtos_britagem[produto_atual][hora] = 1

        # Concentrador
        estoque_pulmao_inicial_concentrador = {}
        for produto, estoque in cenario['concentrador']['estoque_pulmao_inicial_concentrador']:
            estoque_pulmao_inicial_concentrador[produto] = estoque
        min_estoque_pulmao_concentrador = cenario['concentrador']['min_estoque_pulmao_concentrador']
        max_estoque_pulmao_concentrador = cenario['concentrador']['max_estoque_pulmao_concentrador']
        faixas_producao_concentrador = cenario['concentrador']['faixas_producao_concentrador']
        max_taxa_alimentacao = cenario['concentrador']['max_taxa_alimentacao']
        min_taxa_alimentacao = cenario['concentrador']['min_taxa_alimentacao']
        taxa_alimentacao_fixa = cenario['concentrador']['taxa_alimentacao_fixa']
        fixar_taxa_alimentacao = cenario['concentrador']['fixar_taxa_alimentacao']
        numero_faixas_producao = cenario['concentrador']['numero_faixas_producao']

        # Mineroduto

        estoque_eb06_d0 = {"PRDT_C1":0, "PRDT_C2":0, "PRDT_C3":0}
        for produto, estoque in cenario['mineroduto']['estoque_inicial_eb06']:
            estoque_eb06_d0[produto] = estoque

        estoque_eb07_d0 = {}
        for produto, estoque in cenario['mineroduto']['estoque_inicial_eb07']:
            estoque_eb07_d0[produto] = estoque

        estoque_ubu_inicial = {}
        for produto, estoque in cenario['usina']['estoque_inicial_polpa_ubu']:
            estoque_ubu_inicial[produto] = estoque

        bombeamento_matipo = cenario['mineroduto']['bombeamento_matipo']

        min_capacidade_eb_07 = cenario['mineroduto']['min_capacidade_eb07']
        max_capacidade_eb_07 = cenario['mineroduto']['max_capacidade_eb07']

        # Usina
        max_producao_sem_incorporacao = cenario['usina']['max_producao_sem_incorporacao']
        min_producao_sem_incorporacao = cenario['usina']['min_producao_sem_incorporacao']
        producao_sem_incorporacao_fixa = cenario['usina']['producao_sem_incorporacao_fixa']
        fixar_producao_sem_incorporacao = cenario['usina']['fixar_producao_sem_incorporacao']
        min_estoque_polpa_ubu = cenario['usina']['min_estoque_polpa_ubu']
        max_estoque_polpa_ubu = cenario['usina']['max_estoque_polpa_ubu']
        max_taxa_retorno_patio_usina = cenario['usina']['max_taxa_retorno_patio_usina']
        min_estoque_patio_usina = cenario['usina']['min_estoque_patio_usina']
        max_estoque_patio_usina = cenario['usina']['max_estoque_patio_usina']
        estoque_inicial_patio_usina = cenario['usina']['estoque_inicial_patio_usina']

        capacidade_carreg_porto_por_dia = cenario['porto']['capacidade_carreg_porto_por_dia']  # número de navios

        janela_min_bombeamento_polpa = cenario['mineroduto']['janela_min_bombeamento_polpa'] # Mínimo de 1's consecutivos no bombeamento de polpa
        janela_max_bombeamento_polpa = cenario['mineroduto']['janela_max_bombeamento_polpa']
        fixar_janela_bombeamento = cenario['mineroduto']['fixar_janela_bombeamento']
        janela_fixa_bombeamento_polpa = cenario['mineroduto']['janela_fixa_bombeamento_polpa']
        bombeamento_restante_janela_anterior_polpa = cenario['mineroduto']['bombeamento_restante_janela_anterior_polpa']
        bombeamento_restante_janela_anterior_agua = cenario['mineroduto']['bombeamento_restante_janela_anterior_agua']

        janela_min_bombeamento_agua = cenario['mineroduto']['janela_min_bombeamento_agua'] # Mínimo de 1's consecutivos no bombeamento de agua
        janela_max_bombeamento_agua = cenario['mineroduto']['janela_max_bombeamento_agua']
        fixar_janela_bombeamento_agua = cenario['mineroduto']['fixar_janela_bombeamento_agua']
        janela_fixa_bombeamento_agua = cenario['mineroduto']['janela_fixa_bombeamento_agua']
        janela_para_fixar_bombeamento_agua = cenario['mineroduto']['janela_para_fixar_bombeamento_agua']
        nro_janelas_livres_agua = cenario['mineroduto']['nro_janelas_livres_agua']
        janela_livre_min_bombeamento_agua = cenario['mineroduto']['janela_livre_min_bombeamento_agua']
        janela_livre_max_bombeamento_agua = cenario['mineroduto']['janela_livre_max_bombeamento_agua']

        utilizacao_minima_mineroduto = cenario['mineroduto']['utilizacao_minima_mineroduto'] # A quantidade de água bombeada (%) não pode ser maior que 1-param
        janela_da_utilizacao_minima_mineroduto_horas = cenario['mineroduto']['janela_da_utilizacao_minima_mineroduto_horas']

        tempo_mineroduto = cenario['mineroduto']['tempo_mineroduto']

        max_taxa_envio_patio = cenario['mineroduto']['max_taxa_envio_patio']

        fator_limite_excesso_patio = cenario['mineroduto']['fator_limite_excesso_patio']
        vazao_bombas = cenario['mineroduto']['vazao_bombas']
        lim_min_campanha = cenario['mina']['lim_min_campanha']
        lim_max_campanha = cenario['mina']['lim_max_campanha']
        lim_acum_campanha = cenario['mina']['lim_acum_campanha']

        lim_min_janela = cenario['mina']['lim_min_janela']
        lim_max_janela = cenario['mina']['lim_max_janela']
        lim_acum_janela = cenario['mina']['lim_acum_janela']


        # Paradas de manutenção
        # inicio_manutencoes_britagem = cenario['mina']['inicio_manutencoes_britagem']
        # duracao_manutencoes_britagem = cenario['mina']['duracao_manutencoes_britagem']
        #
        # inicio_manutencoes_concentrador = cenario['concentrador']['inicio_manutencoes_concentrador']
        # duracao_manutencoes_concentrador = cenario['concentrador']['duracao_manutencoes_concentrador']
        #
        # inicio_manutencoes_mineroduto = cenario['mineroduto']['inicio_manutencoes_mineroduto']
        # duracao_manutencoes_mineroduto = cenario['mineroduto']['duracao_manutencoes_mineroduto']
        #
        # inicio_manutencoes_usina = cenario['usina']['inicio_manutencoes_usina']
        # duracao_manutencoes_usina = cenario['usina']['duracao_manutencoes_usina']

        def extrair_dia(hora):
            return hora.split('_')[0]

        def extrair_hora(dia):
            return (int(dia[1:])-1)*24

        def indice_da_hora(hora):
            return (int(hora[1:3])-1)*24+int(hora[-2:])-1

        resultados_planilha = {}

        # -----------------------------------------------------------------------------

        print(f'[OK]\nObtendo parâmetros da planilha...   ', end='')

        # Lê os dados da aba MINA
        ws = wb["MINA"]

        # configuração das linhas onde se encontram os parâmetros
        conf_parametros_mina = {
            'Campanha - C3': 2,
            'PPCa - C3': 3,
            'PPCc - C3': 4,
            'Pc - C3': 5,
            'Al2O3a - C3': 6,
            'Al2O3c - C3': 7,
            'Hea - C3': 8,
            'Fe_a - C3': 9,
            'Hec - C3': 10,
            'Umidade - C3': 14,
            'Dif. de Balanço - C3': 15,
            'Sól - C3': 16,
            'DF - C3': 17,
            'UD - C3': 18,
            'Número de Tanque - EB06': 33,
            'Utilização - M3': 35,
            'Disponibilidade - M3': 36,
            'Vazão bombas - M3': 37,
            'Utilização - EB07': 41,
            'Disponibilidade - EB07': 42,
            'Vazão bombas - EB07': 43,
        }

        # colunas onde se encontram os dados da aba mina
        coluna_dia_inicial = 'E'
        coluna_dia_final = 'R'
        coluna_min = 'S'
        coluna_max = 'T'

        # guarda os parâmetros da aba mina
        parametros_mina = {}
        # guarda os limites mínimo e máximo de cada parâmetro
        min_parametros_mina = {}
        max_parametros_mina = {}

        # Lê a aba mina guardando os valores dos parâmetros e os valores mínimos e máximos
        for parametro in conf_parametros_mina:
            linha = conf_parametros_mina[parametro]
            parametros_mina[parametro] = {}
            for idx, cell in enumerate(ws[coluna_dia_inicial+str(linha)+':'+coluna_dia_final+str(linha)][0]):
                parametros_mina[parametro][dias[idx]] = cell.value
                break
            min_parametros_mina[parametro] = ws[coluna_min+str(linha)].value if ws[coluna_min+str(linha)].value is not None else 0
            max_parametros_mina[parametro] = ws[coluna_max+str(linha)].value if ws[coluna_max+str(linha)].value is not None  else BIG_M

        # Lê o parâmetro de geração de lama
        fatorGeracaoLama = ws["D26"].value

        # print('\n CONFERINDO parâmetros da MINA')
        # print(f'{parametros_mina=}')
        # print(f'{min_parametros_mina=}')
        # print(f'{max_parametros_mina=}')
        # print(f'{fatorGeracaoLama=}')

        # Validando se os valores dos parâmetros se encontram dentro dos valores mínimos e máximos

        for parametro in parametros_mina:
            if parametro != 'Campanha - C3':
                if (min(parametros_mina[parametro].values()) < min_parametros_mina[parametro]):
                    print('ATENÇÃO: O valor mínimo de ' + parametro + ' está abaixo do permitido')
                elif (max(parametros_mina[parametro].values()) > max_parametros_mina[parametro]):
                    print('ATENÇÃO: O valor máximo de ' + parametro + ' está acima do permitido')

        # -----------------------------------------------------------------------------

        # Lendo dados da aba MINERODUTO(-D3)

        ws = wb["MINERODUTO(-D3)"]

        # configuração das linhas onde se encontram os parâmetros
        conf_parametros_mineroduto_md3 = {
            'Estoque EB06 -D3': 5,
            'Bombeamento Polpa -D3': 6,
            'Bombeado -D3': 7,
        }

        def get_column_name(column_number):
            column_name = ""
            while column_number > 0:
                column_number, remainder = divmod(column_number - 1, 26)
                column_name = chr(65 + remainder) + column_name
            return column_name

        coluna_hora_inicial_mD3 = 3
        coluna_hora_final_mD3 = coluna_hora_inicial_mD3 + len(horas)*3 - 1

        # guarda os parâmetros da aba MINERODUTO(-D3)
        parametros_mineroduto_md3 = {}

        # Lendo os parâmetros da planilha
        for parametro in conf_parametros_mineroduto_md3:
            linha = conf_parametros_mineroduto_md3[parametro]
            parametros_mineroduto_md3[parametro] = {}
            for idx, cell in enumerate(ws[f'{get_column_name(coluna_hora_inicial_mD3)}{linha}:{get_column_name(coluna_hora_final_mD3)}{linha}'][0]):
                parametros_mineroduto_md3[parametro][horas_Dm3[idx]] = cell.value

        # Sobrescreve o bombeamento polpa -D3 com a informação do cenário, porque agora é multiproduto
        # parametros_mineroduto_md3['Bombeamento Polpa -D3'] = cenario['mineroduto']['bombeamento_polpa_dm3']

        # print('\n CONFERINDO parâmetros da MINERODUTO(-D3)')
        # print(f'{parametros_mineroduto_md3=}')

        # -----------------------------------------------------------------------------

        # Lendo dados da aba MINERODUTO-UBU

        ws = wb["MINERODUTO-UBU"]

        # Lendo parâmetros de uma única célula
        # estoque_eb6_d0 = ws["C5"].value
        #estoque_polpa_ubu = ['estoque_inicial_polpa_ubu']

        estoque_polpa_ubu = cenario['usina']['estoque_inicial_polpa_ubu']
        # configuração das linhas onde se encontram os parâmetros da aba MINERODUTO-UBU
        conf_parametros_mineroduto_ubu = {
            'Capacidade EB06': 7,
            'Polpa Ubu (65Hs) - AJUSTE': 9,
        }

        # colunas onde se encontram os dados da aba MINERODUTO-UBU
        coluna_hora_inicial_d14 = 4
        coluna_hora_final_d14 = coluna_hora_inicial_d14 + len(horas)*len(dias) - 1

        # guarda os parâmetros da aba MINERODUTO-UBU
        parametros_mineroduto_ubu = {}

        # Lendo os parâmetros da planilha
        for parametro in conf_parametros_mineroduto_ubu:
            linha = conf_parametros_mineroduto_ubu[parametro]
            parametros_mineroduto_ubu[parametro] = {}
            for idx, cell in enumerate(ws[f'{get_column_name(coluna_hora_inicial_d14)}{linha}:{get_column_name(coluna_hora_final_d14)}{linha}'][0]):
                parametros_mineroduto_ubu[parametro][horas_D14[idx]] = cell.value

        # print('\n CONFERINDO parâmetros da MINERODUTO-UBU')
        # print(f'{estoque_eb6_d0=}')
        # print(f'{estoque_polpa_ubu=}')
        #print(f'{parametros_mineroduto_ubu=}')

        # -----------------------------------------------------------------------------

        # Lendo dados da aba UBU

        ws = wb["UBU"]

        # configuração das linhas onde se encontram os parâmetros da aba UBU
        conf_parametros_mineroduto_ubu = {
            'SE': 3,
            'PPC': 4,
            '- 325#': 5,
            'Si02': 6,
            'CaO': 7,
            'pH': 8,
            'Densid.': 9,
            # 'H20': 10,
            'UD': 11,
            'DF': 12,
            'Conv.': 13,
        }

        # colunas onde se encontram os dados da aba UBU
        coluna_dia_inicial = 'D'
        coluna_dia_final = 'Q'

        # guarda os parâmetros da aba UBU
        parametros_ubu = {}

        # Lendo os parâmetros da planilha
        for parametro in conf_parametros_mineroduto_ubu:
            linha = conf_parametros_mineroduto_ubu[parametro]
            parametros_ubu[parametro] = {}
            for idx, cell in enumerate(ws[coluna_dia_inicial+str(linha)+':'+coluna_dia_final+str(linha)][0]):
                parametros_ubu[parametro][dias[idx]] = cell.value
                break

        # print('\n CONFERINDO parâmetros da aba UBU')
        # print(f'{parametros_ubu}')

        # -----------------------------------------------------------------------------

        # Lendo dados da aba PÁTIO-PORTO

        ws = wb["PÁTIO-PORTO"]

        # estoque_produto_patio_d0 = ws["D10"].value
        estoque_produto_patio_d0 = cenario['porto']['estoque_produto_patio']

        # -----------------------------------------------------------------------------

        # Lendo dados da aba NAVIOS

        ws = wb["NAVIOS"]

        # configuração das linhas onde se encontram os parâmetros da aba NAVIOS
        conf_parametros_navios = {
            'NAVIOS':'A',
            'DATA-PLANEJADA':'B',
            'DATA-REAL':'C',
            'VOLUME':'D',
        }

        # guarda os parâmetros da aba NAVIOS
        parametros_navios = {}

        # Lendo os parâmetros da planilha
        # Obs.: aqui os navios são indexados por índice, mas isso será alterado mais adiante
        #       porque o mesmo nome de navio pode aparecer mais de uma vez
        for parametro in conf_parametros_navios:
            coluna = conf_parametros_navios[parametro]
            parametros_navios[parametro] = {}
            linha = 2
            cell = ws[coluna+str(linha)]
            while cell.value is not None:
                parametros_navios[parametro][linha-2] = cell.value
                linha += 1
                cell = ws[coluna+str(linha)]

        # -----------------------------------------------------------------------------

        # Guarda os índices dos navios
        navios = cenario['porto']['navios']

        # PORTO

        produtos_de_cada_navio = {navio:{produto_usina:0 for produto_usina in produtos_usina} for navio in navios}
        for navio, produto_usina in cenario['porto']['produtos_de_cada_navio']:
            produtos_de_cada_navio[navio][produto_usina] = 1

        print(f'[OK]\nDefinindo parâmetros calculados...   ', end='')

        # Guarda os parâmetros calculados
        parametros_calculados = {}


        parametros_calculados['RP (Recuperação Mássica) - C3'] = {}
        for dia in dias:
            parametros_calculados['RP (Recuperação Mássica) - C3'][dia] = cenario['mina']['RP'][dia]

        parametros_calculados['Rendimento Operacional - C3'] = {}
        for dia in dias:
            parametros_calculados['Rendimento Operacional - C3'][dia] = cenario['mina']['UD'][dia] * cenario['mina']['DF'][dia]

        parametros_calculados['% Sólidos - EB06'] = {}
        parametros_calculados['Densidade Polpa - EB06'] = {}
        parametros_calculados['Relação: tms - EB06'] = {}
        parametros_calculados['Densidade Polpa - EB07'] = {}
        for dia in dias:
            parametros_calculados['% Sólidos - EB06'][dia] = cenario['mina']['SOL'][dia] - 0.012
            parametros_calculados['Densidade Polpa - EB06'][dia] = 1/((cenario['mina']['SOL_EB06'][dia]/4.75)+(1- cenario['mina']['SOL_EB06'][dia]))
            parametros_calculados['Relação: tms - EB06'][dia] = 5395* cenario['mina']['SOL_EB06'][dia]* cenario['mina']['densidade'][dia]/100
            parametros_calculados['Densidade Polpa - EB07'][dia] =  cenario['mina']['densidade'][dia]
        
        perc_solidos = cenario['mina']['SOL']
        densidade = cenario['mina']['densidade']
        DF = cenario['mina']['DF']
        UD = cenario['mina']['UD']
        umidade = cenario['mina']['umidade']
        RP = cenario['mina']['RP']
        dif_balanco = cenario['mina']['dif_balanco']


        navios_ate_d14 = []
        min_producao_produtos_ubu = cenario['usina']['prod_minima_usina']
        capacidade_patio_porto_min = cenario['porto']['capacidade_patio_porto_min']
        AguaLi = cenario['mineroduto']['janela_min_bombeamento_agua']
        AguaLs = cenario['mineroduto']['janela_max_bombeamento_agua']
        PolpaLi = cenario['mineroduto']['janela_min_bombeamento_polpa']
        PolpaLs = cenario['mineroduto']['janela_max_bombeamento_polpa']
        bomb_polpa_acum_semana_anterior = cenario['mineroduto']['bombeamento_polpa_acum_semana_anterior']
        bomb_agua_acum_semana_anterior = cenario['mineroduto']['bombeamento_agua_acum_semana_anterior']        
        carga_navios = cenario['porto']['carga_navios']
        taxa_carreg_navios = cenario['porto']['taxa_carreg_navios']
        estoque_produto_patio = cenario['porto']['estoque_produto_patio']
        capacidade_carreg_porto_por_dia = cenario['porto']['capacidade_carreg_porto_por_dia']
        capacidade_patio_porto_min = cenario['porto']['capacidade_patio_porto_min']
        capacidade_patio_porto_max = cenario['porto']['capacidade_patio_porto_max']
        data_chegada_navio = cenario['porto']['data_chegada_navio']
        max_capacidade_eb06 = cenario['mineroduto']['max_capacidade_eb06']
        tempo_germano_matipo = cenario['mineroduto']['tempo_germano_matipo']
        tempo_germano_ubu = cenario['mineroduto']['tempo_germano_matipo']
        prod_bomb_hora_anterior = cenario['mineroduto']['prod_polpa_hora_anterior']
        fator_conv = cenario['usina']['fator_conv']
        prod_polpa_hora_anterior = cenario['mineroduto']['prod_polpa_hora_anterior']

        data = {'horas_D14': horas_D14, 'produtos_conc': produtos_conc, 'horas_Dm3_D14': horas_Dm3_D14, 'de_para_produtos_mina_conc': de_para_produtos_mina_conc,
                'min_estoque_pulmao_concentrador': min_estoque_pulmao_concentrador, 'max_estoque_pulmao_concentrador': max_estoque_pulmao_concentrador,
                'numero_faixas_producao': numero_faixas_producao, 'max_taxa_alimentacao': max_taxa_alimentacao, 'parametros_mina': parametros_mina,
                'taxa_producao_britagem': taxa_producao_britagem, 'produtos_britagem': produtos_britagem, 'produtos_mina': produtos_mina,
                'faixas_producao_concentrador': faixas_producao_concentrador, 'estoque_pulmao_inicial_concentrador': estoque_pulmao_inicial_concentrador,
                'parametros_calculados': parametros_calculados, 'fatorGeracaoLama': fatorGeracaoLama, 'parametros_mineroduto_ubu': parametros_mineroduto_ubu,
                'estoque_eb06_d0': estoque_eb06_d0, 'dias': dias, 'args': args, 'max_producao_sem_incorporacao': max_producao_sem_incorporacao,
                'produtos_usina': produtos_usina, 'de_para_produtos_conc_usina': de_para_produtos_conc_usina, 'parametros_ubu': parametros_ubu,
                'tempo_mineroduto': tempo_mineroduto, 'min_estoque_polpa_ubu': min_estoque_polpa_ubu, 'max_estoque_polpa_ubu': max_estoque_polpa_ubu,
                'max_estoque_polpa_ubu': max_estoque_polpa_ubu, 'max_taxa_envio_patio': max_taxa_envio_patio, 'max_taxa_retorno_patio_usina': max_taxa_retorno_patio_usina,
                'min_estoque_patio_usina': min_estoque_patio_usina, 'max_estoque_patio_usina': max_estoque_patio_usina, 'estoque_polpa_ubu': estoque_ubu_inicial,
                'estoque_inicial_patio_usina': estoque_inicial_patio_usina, 'fator_limite_excesso_patio': fator_limite_excesso_patio,
                'capacidade_carreg__por_dia': capacidade_carreg_porto_por_dia, 'navios_ate_d14': navios_ate_d14,
                'horas_Dm3': horas_Dm3, 'navios': navios, 'vazao_bombas': vazao_bombas, 'lim_min_campanha':lim_min_campanha,
                'lim_max_campanha': lim_max_campanha, 'lim_acum_campanha': lim_acum_campanha, 'lim_min_janela': lim_min_janela, 'lim_max_janela':lim_max_janela,
                'lim_acum_janela':lim_acum_janela, 'AguaLi': AguaLi, 'AguaLs': AguaLs, 'PolpaLi': PolpaLi, 'PolpaLs': PolpaLs, 'carga_navios': carga_navios, 
                'taxa_carreg_navios': taxa_carreg_navios, 'estoque_produto_patio': estoque_produto_patio, 'capacidade_carreg_porto_por_dia': capacidade_carreg_porto_por_dia,
                'produtos_navio': produtos_de_cada_navio, 'capacidade_patio_porto_min': capacidade_patio_porto_min, 'capacidade_patio_porto_max': capacidade_patio_porto_max,
                'data_chegada_navio': data_chegada_navio, 'perc_solidos': perc_solidos, 'densidade': densidade, 'DF': DF, 'UD': UD, 'umidade': umidade, 'RP': RP,
                'dif_balanco': dif_balanco, 'bomb_polpa_acum_semana_anterior': bomb_polpa_acum_semana_anterior, 'bomb_agua_acum_semana_anterior': bomb_agua_acum_semana_anterior,
                'max_capacidade_eb06': max_capacidade_eb06, 'tempo_germano_matipo': tempo_germano_matipo, 'tempo_germano_ubu': tempo_germano_ubu, 'prod_bomb_hora_anterior':prod_bomb_hora_anterior,
                'fator_conv': fator_conv, 'prod_polpa_hora_anterior': prod_polpa_hora_anterior, 'min_producao_produtos_ubu': min_producao_produtos_ubu
                }

        return cenario, solver, data

    def load_simplified_data_ppo(self):
        parser = argparse.ArgumentParser(description='Otimizador Plano Semanal')
        parser.add_argument('-c', '--cenario', default='cenarios/ws1.yaml', type=str, help='Caminho para o arquivo do cenário a ser experimentado')
        parser.add_argument('-s', '--solver', default='GUROBI_CMD', type=str, help='Nome do otimizador a ser usado')

        args = parser.parse_args()

        print(f'[OK]\nLendo arquivo {args.cenario}...   ', end='')
        # Abre o arquivo YAML com dados do cenário (parâmetros do problema)
        cenario = self.ler_cenario(args.cenario)

        estoque_eb06_inicial = {}
        for produto, estoque in cenario['mineroduto']['estoque_inicial_eb06']:
            estoque_eb06_inicial[produto] = estoque

        estoque_ubu_inicial = {}
        for produto, estoque in cenario['usina']['estoque_inicial_polpa_ubu']:
            estoque_ubu_inicial[produto] = estoque

        disp_conc_inicial = [10000]*24
        disp_usina_inicial = [10000]*24

        MaxE06 = cenario['mineroduto']['max_capacidade_eb06']
        MaxEUBU = cenario['usina']['max_estoque_polpa_ubu']
        AguaLi = cenario['mineroduto']['janela_min_bombeamento_agua']
        AguaLs = cenario['mineroduto']['janela_max_bombeamento_agua']
        PolpaLi = cenario['mineroduto']['janela_min_bombeamento_polpa']
        PolpaLs = cenario['mineroduto']['janela_max_bombeamento_polpa']
        vazao_bombas = cenario['mineroduto']['vazao_bombas']
        produtos_conc = cenario['concentrador']['produtos_conc']
        produtos_usina = cenario['usina']['produtos_usina']

        return estoque_eb06_inicial, estoque_ubu_inicial, disp_conc_inicial, disp_usina_inicial, MaxE06, MaxEUBU, AguaLi, AguaLs, PolpaLi, PolpaLs, vazao_bombas, produtos_conc, produtos_usina